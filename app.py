# -*- coding: utf-8 -*-
import streamlit as st
import streamlit.components.v1 as components
import json
import urllib.request
import urllib.error
from pathlib import Path
from datetime import datetime

st.set_page_config(
    page_title="오프라인 영업부 달성률",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  .block-container { padding-top: 1.5rem; }
  .stTabs [data-baseweb="tab-list"] { gap:0; width:100%; }
  .stTabs [data-baseweb="tab"] { flex:1; justify-content:center; font-size:1.05rem; font-weight:600; padding:10px 4px; white-space:nowrap; }
  .stTabs [data-baseweb="tab-highlight"] { height:3px; }
</style>
""", unsafe_allow_html=True)

# ─── 타겟 데이터 (2026년 오프라인 영업부) ─────────────────────────────────────────
TARGETS = {
    "TU": {
        "BS VÀ HMP CŨ": {1:860000000,2:860000000,3:860000000,4:860000000,5:860000000,6:860000000,7:860000000,8:860000000,9:860000000,10:860000000,11:860000000,12:860000000},
        "GIẶT XẢ":      {1:171000000,2:171000000,3:171000000,4:171000000,5:171000000,6:171000000,7:None,8:None,9:None,10:None,11:None,12:None},
        "PPSU":         {1:362200000,2:362200000,3:362200000,4:362200000,5:362200000,6:362200000,7:362200000,8:362200000,9:362200000,10:362200000,11:362200000,12:362200000},
        "KHĂN ƯỚT":    {1:1136600000,2:1136600000,3:1136600000,4:1136600000,5:1136600000,6:1136600000,7:1136600000,8:1136600000,9:1136600000,10:1136600000,11:1136600000,12:1136600000},
        "SỮA TẮM":     {1:910000000,2:910000000,3:910000000,4:910000000,5:910000000,6:910000000,7:910000000,8:910000000,9:910000000,10:910000000,11:910000000,12:910000000},
    },
    "VINH": {
        "BS VÀ HMP CŨ": {1:996000000,2:996000000,3:996000000,4:996000000,5:996000000,6:996000000,7:996000000,8:996000000,9:996000000,10:996000000,11:996000000,12:996000000},
        "GIẶT XẢ":      {1:148000000,2:148000000,3:148000000,4:148000000,5:148000000,6:148000000,7:None,8:None,9:None,10:None,11:None,12:None},
        "PPSU":         {1:301000000,2:301000000,3:301000000,4:301000000,5:301000000,6:301000000,7:301000000,8:301000000,9:301000000,10:301000000,11:301000000,12:301000000},
        "KHĂN ƯỚT":    {1:1187000000,2:1187000000,3:1187000000,4:1187000000,5:1187000000,6:1187000000,7:1187000000,8:1187000000,9:1187000000,10:1187000000,11:1187000000,12:1187000000},
        "SỮA TẮM":     {1:1078000000,2:1078000000,3:1078000000,4:1078000000,5:1078000000,6:1078000000,7:1078000000,8:1078000000,9:1078000000,10:1078000000,11:1078000000,12:1078000000},
    },
    "TU,HOI": {
        "BS VÀ HMP CŨ": {1:905000000,2:905000000,3:905000000,4:905000000,5:905000000,6:905000000,7:905000000,8:905000000,9:905000000,10:905000000,11:905000000,12:905000000},
        "GIẶT XẢ":      {1:245000000,2:245000000,3:245000000,4:245000000,5:245000000,6:245000000,7:None,8:None,9:None,10:None,11:None,12:None},
        "PPSU":         {1:695000000,2:695000000,3:695000000,4:695000000,5:695000000,6:695000000,7:695000000,8:695000000,9:695000000,10:695000000,11:695000000,12:695000000},
        "KHĂN ƯỚT":    {1:1560000000,2:1560000000,3:1560000000,4:1560000000,5:1560000000,6:1560000000,7:1560000000,8:1560000000,9:1560000000,10:1560000000,11:1560000000,12:1560000000},
        "SỮA TẮM":     {1:1205000000,2:1205000000,3:1205000000,4:1205000000,5:1205000000,6:1205000000,7:1205000000,8:1205000000,9:1205000000,10:1205000000,11:1205000000,12:1205000000},
    },
    "LAM": {
        "BS VÀ HMP CŨ": {1:1080000000,2:1080000000,3:1080000000,4:1080000000,5:1080000000,6:1080000000,7:1080000000,8:1080000000,9:1080000000,10:1080000000,11:1080000000,12:1080000000},
        "GIẶT XẢ":      {1:240000000,2:240000000,3:240000000,4:240000000,5:240000000,6:240000000,7:None,8:None,9:None,10:None,11:None,12:None},
        "PPSU":         {1:600000000,2:600000000,3:600000000,4:600000000,5:600000000,6:600000000,7:600000000,8:600000000,9:600000000,10:600000000,11:600000000,12:600000000},
        "KHĂN ƯỚT":    {1:2480000000,2:2480000000,3:2480000000,4:2480000000,5:2480000000,6:2480000000,7:2480000000,8:2480000000,9:2480000000,10:2480000000,11:2480000000,12:2480000000},
        "SỮA TẮM":     {1:1380000000,2:1380000000,3:1380000000,4:1380000000,5:1380000000,6:1380000000,7:1380000000,8:1380000000,9:1380000000,10:1380000000,11:1380000000,12:1380000000},
    },
    "HAI": {
        "BS VÀ HMP CŨ": {1:845000000,2:845000000,3:845000000,4:865000000,5:865000000,6:865000000,7:865000000,8:865000000,9:865000000,10:865000000,11:865000000,12:865000000},
        "GIẶT XẢ":      {1:190000000,2:190000000,3:190000000,4:200000000,5:200000000,6:200000000,7:None,8:None,9:None,10:None,11:None,12:None},
        "PPSU":         {1:450000000,2:450000000,3:450000000,4:470000000,5:470000000,6:470000000,7:470000000,8:470000000,9:470000000,10:470000000,11:470000000,12:470000000},
        "KHĂN ƯỚT":    {1:2164000000,2:2164000000,3:2164000000,4:2224000000,5:2224000000,6:2224000000,7:2224000000,8:2224000000,9:2224000000,10:2224000000,11:2224000000,12:2224000000},
        "SỮA TẮM":     {1:935000000,2:935000000,3:935000000,4:945000000,5:945000000,6:945000000,7:945000000,8:945000000,9:945000000,10:945000000,11:945000000,12:945000000},
    },
    "QUOC": {
        "BS VÀ HMP CŨ": {1:1455000000,2:1455000000,3:1455000000,4:1285000000,5:1285000000,6:1285000000,7:1285000000,8:1285000000,9:1285000000,10:1285000000,11:1285000000,12:1285000000},
        "GIẶT XẢ":      {1:265000000,2:265000000,3:265000000,4:235000000,5:235000000,6:235000000,7:None,8:None,9:None,10:None,11:None,12:None},
        "PPSU":         {1:795000000,2:795000000,3:795000000,4:705000000,5:705000000,6:705000000,7:705000000,8:705000000,9:705000000,10:705000000,11:705000000,12:705000000},
        "KHĂN ƯỚT":    {1:2600000000,2:2600000000,3:2600000000,4:2530000000,5:2530000000,6:2530000000,7:2530000000,8:2530000000,9:2530000000,10:2530000000,11:2530000000,12:2530000000},
        "SỮA TẮM":     {1:1650000000,2:1650000000,3:1650000000,4:1450000000,5:1450000000,6:1450000000,7:1450000000,8:1450000000,9:1450000000,10:1450000000,11:1450000000,12:1450000000},
    },
    "HUNG": {
        "BS VÀ HMP CŨ": {1:1418000000,2:1418000000,3:1418000000,4:1418000000,5:1418000000,6:1418000000,7:1418000000,8:1418000000,9:1418000000,10:1418000000,11:1418000000,12:1418000000},
        "GIẶT XẢ":      {1:222750000,2:222750000,3:222750000,4:222750000,5:222750000,6:222750000,7:None,8:None,9:None,10:None,11:None,12:None},
        "PPSU":         {1:813000000,2:813000000,3:813000000,4:813000000,5:813000000,6:813000000,7:813000000,8:813000000,9:813000000,10:813000000,11:813000000,12:813000000},
        "KHĂN ƯỚT":    {1:1896000000,2:1896000000,3:1896000000,4:1896000000,5:1896000000,6:1896000000,7:1896000000,8:1896000000,9:1896000000,10:1896000000,11:1896000000,12:1896000000},
        "SỮA TẮM":     {1:879500000,2:879500000,3:879500000,4:879500000,5:879500000,6:879500000,7:879500000,8:879500000,9:879500000,10:879500000,11:879500000,12:879500000},
    },
    "NHU": {
        "BS VÀ HMP CŨ": {1:1821000000,2:1821000000,3:1821000000,4:1821000000,5:1821000000,6:1821000000,7:1821000000,8:1821000000,9:1821000000,10:1821000000,11:1821000000,12:1821000000},
        "GIẶT XẢ":      {1:360000000,2:360000000,3:360000000,4:360000000,5:360000000,6:360000000,7:None,8:None,9:None,10:None,11:None,12:None},
        "PPSU":         {1:1003000000,2:1003000000,3:1003000000,4:1003000000,5:1003000000,6:1003000000,7:1003000000,8:1003000000,9:1003000000,10:1003000000,11:1003000000,12:1003000000},
        "KHĂN ƯỚT":    {1:2140000000,2:2140000000,3:2140000000,4:2140000000,5:2140000000,6:2140000000,7:2140000000,8:2140000000,9:2140000000,10:2140000000,11:2140000000,12:2140000000},
        "SỮA TẮM":     {1:1177000000,2:1177000000,3:1177000000,4:1177000000,5:1177000000,6:1177000000,7:1177000000,8:1177000000,9:1177000000,10:1177000000,11:1177000000,12:1177000000},
    },
    "VAN": {
        "BS VÀ HMP CŨ": {1:2422000000,2:2422000000,3:2422000000,4:2882000000,5:2882000000,6:2882000000,7:3498000000,8:3498000000,9:3498000000,10:4238000000,11:4238000000,12:4238000000},
        "GIẶT XẢ":      {1:50000000,2:50000000,3:50000000,4:50000000,5:50000000,6:50000000,7:None,8:None,9:None,10:None,11:None,12:None},
        "PPSU":         {1:533000000,2:533000000,3:533000000,4:638000000,5:638000000,6:638000000,7:757000000,8:757000000,9:757000000,10:1007000000,11:1007000000,12:1007000000},
        "KHĂN ƯỚT":    {1:1028000000,2:1028000000,3:1028000000,4:1028000000,5:1028000000,6:1028000000,7:1028000000,8:1028000000,9:1028000000,10:1028000000,11:1028000000,12:1028000000},
        "SỮA TẮM":     {1:623000000,2:623000000,3:623000000,4:743000000,5:743000000,6:743000000,7:990000000,8:990000000,9:990000000,10:1295000000,11:1295000000,12:1295000000},
    },
    "WINMART": {
        "BS VÀ HMP CŨ": {m:0 for m in range(1,13)},
        "GIẶT XẢ":      {m:0 for m in range(1,13)},
        "PPSU":         {m:0 for m in range(1,13)},
        "KHĂN ƯỚT":    {1:850000000,2:850000000,3:850000000,4:850000000,5:850000000,6:850000000,7:850000000,8:850000000,9:850000000,10:850000000,11:850000000,12:850000000},
        "SỮA TẮM":     {m:0 for m in range(1,13)},
    },
    "LOTTEMART": {
        "BS VÀ HMP CŨ": {m:0 for m in range(1,13)},
        "GIẶT XẢ":      {m:0 for m in range(1,13)},
        "PPSU":         {m:0 for m in range(1,13)},
        "KHĂN ƯỚT":    {1:1500000000,2:1500000000,3:1500000000,4:1600000000,5:1600000000,6:1600000000,7:1600000000,8:1600000000,9:1600000000,10:1600000000,11:1600000000,12:1600000000},
        "SỮA TẮM":     {m:0 for m in range(1,13)},
    },
}

SKU_LIST  = ["BS VÀ HMP CŨ", "GIẶT XẢ", "PPSU", "KHĂN ƯỚT", "SỮA TẮM"]
ASM_LIST  = ["TU", "VINH", "TU,HOI", "LAM", "HAI", "QUOC", "HUNG", "NHU", "VAN", "WINMART", "LOTTEMART"]
SKU_ICON  = {"BS VÀ HMP CŨ": "🍼", "GIẶT XẢ": "👕", "PPSU": "👑", "KHĂN ƯỚT": "🧻", "SỮA TẮM": "🛁"}
DATA_DIR  = Path(__file__).parent / "data"
DATA_DIR.mkdir(exist_ok=True)
GIST_FILE = "offline_records.json"


# ─── Gist 헬퍼 ──────────────────────────────────────────────────────────────────
def _use_gist():
    try:
        return bool(st.secrets.get("GIST_TOKEN") and st.secrets.get("GIST_ID"))
    except Exception:
        return False

def _gist_req(method, data=None):
    token   = st.secrets["GIST_TOKEN"]
    gist_id = st.secrets["GIST_ID"]
    body    = json.dumps(data).encode() if data else None
    req     = urllib.request.Request(
        f"https://api.github.com/gists/{gist_id}", data=body, method=method,
        headers={"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json",
                 "Content-Type": "application/json"})
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def load_records():
    if _use_gist():
        try:
            res = _gist_req("GET")
            files = res.get("files", {})
            if GIST_FILE in files:
                return json.loads(files[GIST_FILE]["content"])
        except Exception:
            pass
    local = DATA_DIR / "offline_records.json"
    return json.loads(local.read_text(encoding="utf-8")) if local.exists() else {}

def save_records(records):
    if _use_gist():
        _gist_req("PATCH", {"files": {GIST_FILE: {"content": json.dumps(records, ensure_ascii=False, indent=2)}}})
    else:
        (DATA_DIR / "offline_records.json").write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")


# ─── 유틸 ──────────────────────────────────────────────────────────────────────
def fmt_b(v):
    if v is None or v == 0:
        return "0.00 tỷ"
    return f"{v/1e9:.2f} tỷ"

def get_target(asm, sku, month):
    return TARGETS.get(asm, {}).get(sku, {}).get(month)

def get_actual(records, month, asm, sku):
    return records.get(str(month), {}).get(asm, {}).get(sku, 0) or 0

def month_summary(records, month):
    """월 전체 합계: (total_target, total_actual, under_count)"""
    total_t, total_a, under = 0, 0, 0
    for sku in SKU_LIST:
        st_val, sa_val = 0, 0
        for asm in ASM_LIST:
            t = get_target(asm, sku, month)
            if t is None or t == 0:
                continue
            st_val += t
            sa_val += get_actual(records, month, asm, sku)
        if st_val > 0:
            total_t += st_val
            total_a += sa_val
            if sa_val < st_val:
                under += 1
    return total_t, total_a, under

def month_icon(records, month):
    if str(month) not in records:
        return "⬜"
    t, a, _ = month_summary(records, month)
    if t == 0:
        return "⬜"
    p = a / t * 100
    return "✅" if p >= 100 else "🟡" if p >= 70 else "🔴"


# ─── Excel 파싱 ─────────────────────────────────────────────────────────────────
def parse_excel_actuals(file_bytes):
    import openpyxl, re
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # 월/연도 감지
    month = year = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for cell in row:
            if cell and 'K' in str(cell) and '/' in str(cell):
                m = re.search(r'(\d{1,2})/(\d{4})', str(cell))
                if m:
                    month, year = int(m.group(1)), int(m.group(2))
                    break
        if month:
            break
    if not month:
        raise ValueError("파일에서 월 정보(Kỳ)를 찾을 수 없습니다")

    # 실적 컬럼 인덱스 (0-based)
    SKU_COLS = {
        "BS VÀ HMP CŨ": [14],
        "GIẶT XẢ":      [16],
        "PPSU":         [20],
        "KHĂN ƯỚT":    [22, 26],
        "SỮA TẮM":     [24],
    }

    def get_val(row, cols):
        return sum((row[c] or 0) for c in cols if c < len(row))

    result = {asm: {sku: 0 for sku in SKU_LIST} for asm in ASM_LIST}

    # ASM 코드 → 대시보드 ASM명
    ASM_CODE = {
        "02401": "TU",
        "03344": "VINH",
        "00887": "LAM",
        "01683": "HAI",
        "00331": "QUOC",
        "00318": "HUNG",
        "01987": "NHU",
        "01565": "VAN",
    }
    if month <= 6:
        ASM_CODE["00323"] = "TU,HOI"

    for row in ws.iter_rows(min_row=11, values_only=True):
        if not row[0] or not row[1]:
            continue
        code  = str(row[0]).strip()
        level = str(row[1]).strip()
        name  = str(row[2]).strip() if row[2] else ''

        # ASM 합계 행
        if level == 'ASM' and name.startswith('Total - '):
            asm = ASM_CODE.get(code)
            if asm:
                for sku, cols in SKU_COLS.items():
                    result[asm][sku] += get_val(row, cols)
            # TU,HOI 7월 이후: TU(02401) 포함
            if month >= 7 and code == "02401":
                for sku, cols in SKU_COLS.items():
                    result["TU,HOI"][sku] += get_val(row, cols)

        # TU,HOI 7월 이후: HOI(02191) SUP 합계 포함
        if level == 'Total - SUP' and month >= 7 and code == "02191":
            for sku, cols in SKU_COLS.items():
                result["TU,HOI"][sku] += get_val(row, cols)

        # WINMART / LOTTEMART (고객명 필터, 개별 행)
        if row[4] and level not in ('Total - SUP', 'ASM'):
            customer = str(row[4]).upper()
            ch = None
            if 'WINMART' in customer:
                ch = 'WINMART'
            elif 'LOTTE' in customer:
                ch = 'LOTTEMART'
            if ch:
                for sku, cols in SKU_COLS.items():
                    result[ch][sku] += get_val(row, cols)

    return month, year, result


# ─── HTML 렌더러 ────────────────────────────────────────────────────────────────
def bar_colors(pct):
    if pct >= 100:
        return "#10b981", "linear-gradient(90deg,#059669,#10b981)"
    if pct >= 70:
        return "#f59e0b", "linear-gradient(90deg,#d97706,#f59e0b)"
    return "#ef4444", "linear-gradient(90deg,#dc2626,#ef4444)"

def _sku_bar_row(icon, sku, pct, sku_a, sku_t, solid, grad):
    bar_w = min(pct, 100)
    over = '<span style="background:#d1fae5;color:#065f46;font-size:10px;font-weight:700;padding:2px 7px;border-radius:99px;margin-left:8px;">OVER</span>' if pct > 100 else ""
    return f"""
    <div style="margin-bottom:8px;">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:7px;">
        <div style="display:flex;align-items:center;gap:8px;">
          <span style="font-size:16px;">{icon}</span>
          <span style="font-weight:600;font-size:14px;color:#1e293b;">{sku}</span>{over}
        </div>
        <span style="font-weight:700;font-size:15px;color:{solid};">{pct:.1f}%</span>
      </div>
      <div style="background:#f1f5f9;border-radius:99px;height:10px;overflow:visible;position:relative;">
        <div style="width:{bar_w:.1f}%;background:{grad};height:100%;border-radius:99px;position:relative;">
          <div style="position:absolute;right:-1px;top:50%;transform:translateY(-50%);width:16px;height:16px;background:{solid};border-radius:50%;border:2px solid white;box-shadow:0 0 0 2px {solid}33;"></div>
        </div>
      </div>
      <div style="display:flex;justify-content:space-between;margin-top:5px;font-size:12px;color:#94a3b8;">
        <span>실적 <strong style="color:#475569;">{fmt_b(sku_a)}</strong></span>
        <span>목표 <strong style="color:#475569;">{fmt_b(sku_t)}</strong></span>
      </div>
    </div>"""


def render_dashboard_html(records, month):
    total_t, total_a, under_count = month_summary(records, month)
    total_pct  = total_a / total_t * 100 if total_t else 0
    pct_color  = "#10b981" if total_pct >= 100 else "#f59e0b" if total_pct >= 70 else "#ef4444"
    under_color = "#10b981" if under_count == 0 else "#ef4444"

    # ── 요약 카드 ──
    summary = f"""
    <div class="sg">
      <div class="sc"><div class="sl">전체 달성률</div><div class="sv" style="color:{pct_color};">{total_pct:.1f}%</div></div>
      <div class="sc"><div class="sl">실적 합계</div><div class="sv" style="color:#f1f5f9;">{fmt_b(total_a)}</div></div>
      <div class="sc"><div class="sl">목표 합계</div><div class="sv" style="color:#f1f5f9;">{fmt_b(total_t)}</div></div>
      <div class="sc"><div class="sl">미달 SKU</div><div class="sv" style="color:{under_color};">{under_count}개</div></div>
    </div>"""

    # ── SKU 아코디언 + ASM 드릴다운 ──
    sku_sections = ""
    for idx, sku in enumerate(SKU_LIST):
        icon = SKU_ICON.get(sku, "")
        sku_t, sku_a = 0, 0
        for asm in ASM_LIST:
            t = get_target(asm, sku, month)
            if t is None or t == 0:
                continue
            sku_t += t
            sku_a += get_actual(records, month, asm, sku)

        if sku_t == 0:
            sku_sections += f"""
            <div style="margin-bottom:12px;padding:14px 16px;background:white;border-radius:12px;border:1px solid #e2e8f0;opacity:.45;">
              <div style="display:flex;align-items:center;gap:8px;">
                <span style="font-size:16px;">{icon}</span>
                <span style="font-weight:600;font-size:14px;color:#94a3b8;">{sku}</span>
                <span style="font-size:12px;color:#64748b;margin-left:auto;">단종 / 해당 없음</span>
              </div>
            </div>"""
            continue

        pct = sku_a / sku_t * 100
        solid, grad = bar_colors(pct)
        bar_row = _sku_bar_row(icon, sku, pct, sku_a, sku_t, solid, grad)

        # ASM 드릴다운 행들
        asm_rows = ""
        CHANNEL_ASMS = {"WINMART", "LOTTEMART"}
        for asm in ASM_LIST:
            t = get_target(asm, sku, month)
            if t is None or t == 0:
                continue
            a = get_actual(records, month, asm, sku)
            p = a / t * 100
            as_, ag = bar_colors(p)
            abw = min(p, 100)
            is_channel = asm in CHANNEL_ASMS
            row_bg = "background:#f0f9ff;border-left:3px solid #38bdf8;" if is_channel else ""
            asm_label = f'<span style="font-size:9px;color:#0284c7;background:#e0f2fe;padding:1px 5px;border-radius:4px;margin-left:4px;">채널</span>' if is_channel else ""
            if p >= 100:
                badge = f'<span style="background:#d1fae5;color:#065f46;font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px;white-space:nowrap;">✅ 달성</span>'
            elif p >= 70:
                badge = f'<span style="background:#fef3c7;color:#92400e;font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px;white-space:nowrap;">⚠️ 진행중</span>'
            else:
                badge = f'<span style="background:#fee2e2;color:#991b1b;font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px;white-space:nowrap;">🚨 미달</span>'

            asm_rows += f"""
            <div style="display:grid;grid-template-columns:100px 1fr 90px 90px 60px;align-items:center;gap:10px;padding:10px 0;border-bottom:1px solid #f1f5f9;{row_bg}padding-left:8px;">
              <span style="font-weight:700;font-size:13px;color:{'#0369a1' if is_channel else '#334155'};">{asm}{asm_label}</span>
              <div style="background:#f1f5f9;border-radius:99px;height:8px;position:relative;">
                <div style="width:{abw:.1f}%;background:{ag};height:100%;border-radius:99px;position:relative;">
                  <div style="position:absolute;right:-1px;top:50%;transform:translateY(-50%);width:12px;height:12px;background:{as_};border-radius:50%;border:2px solid white;"></div>
                </div>
              </div>
              <span style="font-size:12px;color:#64748b;text-align:right;">{fmt_b(a)}</span>
              <span style="font-size:12px;color:#94a3b8;text-align:right;">{fmt_b(t)}</span>
              <span style="font-weight:700;font-size:13px;color:{as_};text-align:right;">{p:.1f}%</span>
            </div>"""

        sid = f"sku_{idx}"
        sku_sections += f"""
        <div style="margin-bottom:10px;background:white;border-radius:12px;border:1px solid #e2e8f0;overflow:hidden;">
          <div onclick="toggle('{sid}')" style="padding:16px 18px;cursor:pointer;user-select:none;">
            {bar_row}
            <div style="display:flex;align-items:center;gap:6px;margin-top:8px;font-size:12px;color:#94a3b8;">
              <span id="arrow_{sid}" style="transition:transform .2s;">▶</span>
              <span>ASM별 상세 보기</span>
            </div>
          </div>
          <div id="{sid}" style="display:none;padding:0 18px 16px;border-top:1px solid #f1f5f9;">
            <div style="display:grid;grid-template-columns:80px 1fr 90px 90px 60px;gap:10px;padding:8px 0;font-size:11px;font-weight:600;color:#94a3b8;letter-spacing:.04em;text-transform:uppercase;">
              <span>ASM</span><span>달성 진행</span><span style="text-align:right">실적</span><span style="text-align:right">목표</span><span style="text-align:right">달성률</span>
            </div>
            {asm_rows}
          </div>
        </div>"""

    # 전체 합계 바 (ASM 아코디언)
    tb_w = min(total_pct, 100)
    ts, tg = bar_colors(total_pct)

    asm_total_rows = ""
    for asm in ASM_LIST:
        asm_t, asm_a = 0, 0
        for sku in SKU_LIST:
            t = get_target(asm, sku, month)
            if t is None or t == 0:
                continue
            asm_t += t
            asm_a += get_actual(records, month, asm, sku)
        if asm_t == 0:
            continue
        p = asm_a / asm_t * 100
        as_, ag = bar_colors(p)
        abw = min(p, 100)
        if p >= 100:
            badge = '<span style="background:#d1fae5;color:#065f46;font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px;white-space:nowrap;">✅ 달성</span>'
        elif p >= 70:
            badge = '<span style="background:#fef3c7;color:#92400e;font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px;white-space:nowrap;">⚠️ 진행중</span>'
        else:
            badge = '<span style="background:#fee2e2;color:#991b1b;font-size:10px;font-weight:700;padding:2px 8px;border-radius:99px;white-space:nowrap;">🚨 미달</span>'
        asm_total_rows += f"""
        <div style="display:grid;grid-template-columns:80px 1fr 90px 90px 60px;align-items:center;gap:10px;padding:10px 0;border-bottom:1px solid #f1f5f9;">
          <span style="font-weight:700;font-size:13px;color:#334155;">{asm}</span>
          <div style="background:#f1f5f9;border-radius:99px;height:8px;position:relative;">
            <div style="width:{abw:.1f}%;background:{ag};height:100%;border-radius:99px;position:relative;">
              <div style="position:absolute;right:-1px;top:50%;transform:translateY(-50%);width:12px;height:12px;background:{as_};border-radius:50%;border:2px solid white;"></div>
            </div>
          </div>
          <span style="font-size:12px;color:#64748b;text-align:right;">{fmt_b(asm_a)}</span>
          <span style="font-size:12px;color:#94a3b8;text-align:right;">{fmt_b(asm_t)}</span>
          <span style="font-weight:700;font-size:13px;color:{as_};text-align:right;">{p:.1f}%</span>
        </div>"""

    total_bar = f"""
    <div style="background:white;border-radius:12px;border:1px solid #e2e8f0;overflow:hidden;margin-top:4px;">
      <div onclick="toggle('total_asm')" style="padding:16px 18px;cursor:pointer;user-select:none;">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:7px;">
          <span style="font-weight:700;font-size:15px;color:#1e293b;">📊 전체 합계</span>
          <span style="font-weight:700;font-size:16px;color:{ts};">{total_pct:.1f}%</span>
        </div>
        <div style="background:#f1f5f9;border-radius:99px;height:13px;overflow:visible;position:relative;">
          <div style="width:{tb_w:.1f}%;background:{tg};height:100%;border-radius:99px;position:relative;">
            <div style="position:absolute;right:-1px;top:50%;transform:translateY(-50%);width:19px;height:19px;background:{ts};border-radius:50%;border:2px solid white;box-shadow:0 0 0 3px {ts}33;"></div>
          </div>
        </div>
        <div style="display:flex;justify-content:space-between;margin-top:5px;font-size:12px;color:#94a3b8;">
          <span>실적 <strong style="color:#475569;">{fmt_b(total_a)}</strong></span>
          <span>목표 <strong style="color:#475569;">{fmt_b(total_t)}</strong></span>
        </div>
        <div style="display:flex;align-items:center;gap:6px;margin-top:8px;font-size:12px;color:#94a3b8;">
          <span id="arrow_total_asm" style="transition:transform .2s;">▶</span>
          <span>ASM별 상세 보기</span>
        </div>
      </div>
      <div id="total_asm" style="display:none;padding:0 18px 16px;border-top:1px solid #f1f5f9;">
        <div style="display:grid;grid-template-columns:80px 1fr 90px 90px 60px;gap:10px;padding:8px 0;font-size:11px;font-weight:600;color:#94a3b8;letter-spacing:.04em;text-transform:uppercase;">
          <span>ASM</span><span>달성 진행</span><span style="text-align:right">실적</span><span style="text-align:right">목표</span><span style="text-align:right">달성률</span>
        </div>
        {asm_total_rows}
      </div>
    </div>"""

    return f"""
    <style>
      body{{margin:0;overflow:hidden;}}
      .sg{{display:grid;grid-template-columns:repeat(4,1fr);gap:12px;margin-bottom:16px;}}
      @media(max-width:600px){{.sg{{grid-template-columns:repeat(2,1fr);}}}}
      .sc{{background:#0f172a;border-radius:14px;padding:16px;}}
      .sl{{font-size:11px;font-weight:600;letter-spacing:.06em;color:#64748b;text-transform:uppercase;margin-bottom:8px;}}
      .sv{{font-size:24px;font-weight:700;letter-spacing:-.5px;}}
    </style>
    <script>
      function sendHeight() {{
        var h = document.getElementById('root').scrollHeight + 24;
        try {{
          // 같은 도메인이므로 부모 DOM 직접 조작 가능
          var frames = window.parent.document.querySelectorAll('iframe');
          frames.forEach(function(f) {{
            if (f.contentWindow === window) {{
              f.style.height = h + 'px';
              f.style.minHeight = h + 'px';
            }}
          }});
        }} catch(e) {{}}
      }}
      function toggle(id) {{
        var el = document.getElementById(id);
        var arrow = document.getElementById('arrow_' + id);
        if (el.style.display === 'none') {{
          el.style.display = 'block';
          arrow.style.transform = 'rotate(90deg)';
        }} else {{
          el.style.display = 'none';
          arrow.style.transform = 'rotate(0deg)';
        }}
        sendHeight();
      }}
      window.addEventListener('load', function() {{ sendHeight(); }});
    </script>
    <div id="root" style="font-family:'Inter',system-ui,sans-serif;background:#f8fafc;padding:16px;border-radius:16px;">
      {summary}
      {sku_sections}
      {total_bar}
    </div>"""


# ─── 어드민 입력 ────────────────────────────────────────────────────────────────
def admin_page(records, month):
    st.markdown(f"### ✏️ {month}월 실적 입력")

    tab_manual, tab_excel = st.tabs(["✏️ 수동 입력", "📂 Excel 업로드"])

    with tab_excel:
        st.markdown("**월별 매출 Excel 파일을 업로드하면 자동으로 실적을 불러옵니다.**")
        uploaded = st.file_uploader("Excel 파일 선택", type=["xlsx"], key=f"xl_{month}")
        if uploaded:
            try:
                file_bytes = uploaded.read()
                m, y, data = parse_excel_actuals(file_bytes)
                st.success(f"✅ {y}년 {m}월 데이터 파싱 완료")

                # 미리보기
                preview_rows = []
                for asm in ASM_LIST:
                    row_data = {"ASM": asm}
                    for sku in SKU_LIST:
                        v = data[asm].get(sku, 0)
                        row_data[sku] = f"{v/1e9:.2f}tỷ" if v else "-"
                    preview_rows.append(row_data)
                import pandas as pd
                st.dataframe(pd.DataFrame(preview_rows).set_index("ASM"), use_container_width=True)

                if st.button("💾 이 데이터로 저장", type="primary", key=f"xl_save_{month}"):
                    ms = str(m)
                    records.setdefault(ms, {})
                    for asm in ASM_LIST:
                        records[ms].setdefault(asm, {})
                        for sku in SKU_LIST:
                            v = data[asm].get(sku, 0)
                            if v:
                                records[ms][asm][sku] = v
                    save_records(records)
                    st.success(f"✅ {m}월 실적 저장 완료!")
                    st.rerun()
            except Exception as e:
                st.error(f"파싱 오류: {e}")

    with tab_manual:
        sel_asm = st.selectbox("ASM 선택", ASM_LIST, key=f"asm_sel_{month}")
        st.markdown(f"**{sel_asm}** — {month}월 실적 (단위: VND)")
        st.markdown("---")

        inputs = {}
        for sku in SKU_LIST:
            t = get_target(sel_asm, sku, month)
            if t is None:
                st.markdown(f"- **{SKU_ICON.get(sku,'')} {sku}**: 단종 (7월~)")
                continue
            if t == 0:
                continue
            cur = get_actual(records, month, sel_asm, sku)
            c1, c2 = st.columns([3, 2])
            with c1:
                val = st.number_input(
                    f"{SKU_ICON.get(sku,'')} {sku}  (목표: {fmt_b(t)})",
                    min_value=0, value=int(cur), step=1_000_000,
                    key=f"inp_{sel_asm}_{sku}_{month}", format="%d")
            with c2:
                p = val / t * 100 if t else 0
                solid, _ = bar_colors(p)
                st.markdown(f"<div style='padding-top:28px;font-weight:700;color:{solid}'>{p:.1f}%</div>", unsafe_allow_html=True)
            inputs[sku] = val

        st.markdown("---")
        if st.button("💾 저장", type="primary", use_container_width=True, key=f"save_{month}"):
            ms = str(month)
            records.setdefault(ms, {}).setdefault(sel_asm, {})
            for sku, val in inputs.items():
                records[ms][sel_asm][sku] = val
            try:
                save_records(records)
                st.success(f"✅ {sel_asm} {month}월 실적 저장 완료!")
                st.rerun()
            except Exception as e:
                st.error(f"저장 실패: {e}")


# ─── 메인 ──────────────────────────────────────────────────────────────────────
def main():
    # 인증
    if "role" not in st.session_state:
        st.markdown("<h2 style='text-align:center;margin-top:80px;'>🔐 오프라인 영업부 달성률</h2>", unsafe_allow_html=True)
        c1, c2, c3 = st.columns([1, 2, 1])
        with c2:
            pwd = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요")
            if st.button("로그인", use_container_width=True, type="primary"):
                try:
                    apw = st.secrets.get("ADMIN_PASSWORD", "")
                    vpw = st.secrets.get("APP_PASSWORD", "")
                except Exception:
                    apw = vpw = ""
                if apw and pwd == apw:
                    st.session_state["role"] = "admin"
                    st.rerun()
                elif vpw and pwd == vpw:
                    st.session_state["role"] = "viewer"
                    st.rerun()
                else:
                    st.error("비밀번호가 올바르지 않습니다.")
        return

    role = st.session_state["role"]

    # 사이드바
    with st.sidebar:
        st.markdown("## 📊 오프라인 영업부")
        st.markdown(f"**접속 권한:** {'관리자' if role == 'admin' else '조회'}")
        if role == "admin":
            st.markdown("---")
            page = st.radio("화면", ["📊 대시보드", "✏️ 실적 입력"])
        else:
            page = "📊 대시보드"
        st.markdown("---")
        if st.button("🚪 로그아웃"):
            del st.session_state["role"]
            st.rerun()

    records = load_records()
    st.title("📊 오프라인 영업부 달성 현황")

    # ── 월별 기록 버튼 ──
    st.markdown("### 📅 월별 기록")
    components.html("""<script>
var s = window.parent.document.createElement('style');
s.textContent = 'button[data-testid="stBaseButton-secondary"] p { white-space: nowrap !important; font-size: 12px !important; }';
window.parent.document.head.appendChild(s);
</script>""", height=0)
    month_cols = st.columns(12)
    for m, col in zip(range(1, 13), month_cols):
        with col:
            icon = month_icon(records, m)
            label = f"{icon} {m}월"
            if st.button(label, key=f"mbtn_{m}", use_container_width=True):
                st.session_state["view_month"] = m
                st.session_state.pop("sel_sku", None)

    st.markdown("---")

    view_month = st.session_state.get("view_month")
    if not view_month:
        st.info("위 월별 버튼을 클릭하면 달성 현황이 표시됩니다.")
        return

    st.subheader(f"📈 {view_month}월 달성 현황")

    if page == "✏️ 실적 입력":
        admin_page(records, view_month)
        return

    # ── 대시보드 HTML (아코디언 포함, 높이 자동조절) ──
    html = render_dashboard_html(records, view_month)
    components.html(html, height=700, scrolling=False)


if __name__ == "__main__":
    main()
